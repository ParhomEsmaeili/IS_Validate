import argparse
import json
import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from src.general_utils.dict_utils import extractor, dict_path_modif
from typing import Dict, List, Any, Tuple
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RESULT_SEARCHSTRING = {
    'Dice': 'Dice',
    'NSD': 'NSD',
    'NoI': ['NOI', 'Failure_Cases_Fraction']
}
MAPPING_SUBCATEGORIES = {
    'auc': 'nAUC',
    'Init.': 'Init.',
    'Interactive Edit Iter': 'Interactive Edit Iter',
    'NOI': 'nNoI',
    'Failure_Cases_Fraction': 'NoF'
}

TABLE_MAP = {
    'Dice_median': '',
    'NSD_median': '',
    'Dice_auc_median': 'nAUC',
    'NSD_auc_median': 'nAUC',
    'Normalised_Median_NOI': 'nNoI',
    'Interactive Edit Iter': 'Iter.',
    'Failure_Cases_Fraction': 'NoF'
}

COLUMN_ORDERING = [
    'Init.',
    'Interactive Edit Iter',
    'nAUC',
    'nNoI',
    'NoF'
]

DATASET_MAPPING = {
    'Dataset001_BrainTumour': 'Brain Tumour Core',
    'Dataset002_Heart': 'Heart',
    'Dataset003_Liver': 'Whole Liver',
    'Dataset004_Hippocampus': 'Whole Hippocampus',
    'Dataset005_Prostate': 'Whole Prostate',
    'Dataset006_Lung': 'Whole Lung',
    'Dataset007_Pancreas': 'Whole Pancreas',
    'Dataset008_HepaticVessels': 'Hepatic Vessels',
    'Dataset009_Spleen': 'Spleen',
    'Dataset010_Colon': 'Colon'
}

ALGORITHM_MAPPING = {
    'sam2v1': 'SAM2',
    'sammed2dv1': 'SAM-Med2D',
    'sammed3dv1': 'SAM-Med3D',
    'segvolv1': 'SegVol',
    'nnintv1': 'nnInteractive'
}
#!/usr/bin/env python3
"""
table_generator.py

Generate a summary table (Excel + LaTeX) from a list of algorithm folders.

metric_config.json format (example):
{
  "metric_type_1_path.csv": ["metric_type_1_column_1", "metric_type_1_column_2"],
  "metric_type_2_path.json": ["metric_type_2_column_2", "metric_type_2_column_2"]
}  
Values of the dictionary indicate the columns which need to be extracted for printing, and they also serve as descriptors for
the metrics too when generating tables.
"""

def filter_string_table(input_str: str):
    for key, val in TABLE_MAP.items():
        if key in input_str:
            input_str = input_str.replace(key, val)
    return input_str 

def load_metrics_config(json_dict:str) -> Dict[str, List[str]]:
    cfg = json.loads(json_dict)
    if not isinstance(cfg, dict):
        raise ValueError("metrics config must be a JSON object mapping metric -> [metric information]")
    return cfg


# Example usage:
# my_dict = {'Dice_ ': 1, 'NSD_  ': 2, 'Other': 3}
# my_dict = regex_replace_dict_keys(my_dict)
def sort_by_substring_order(input_list: List[str], priority_list: List[str]) -> List[str]:
    def sort_key(input_list):
        for i, substr in enumerate(priority_list):
            if substr.lower() in input_list.lower():
                return i
        return len(input_list)  # put unmatched at the end
    return sorted(input_list, key=sort_key)

def parse_csv_file(path: str) -> Dict[str, Any]:
    """
    Attempt to parse a CSV of metrics. We support two common layouts:
      - Single row with columns being metric names
      - Two columns: metric_name, value (multiple rows)
    Returns a flat dict metric_name -> value
    """
    try:
        df = pd.read_csv(path)
    except Exception:
        return {}
    if df.shape[0] == 1:
        # use column names
        row = df.iloc[0]
        return {str(col): row[col] for col in df.columns}
    # else, if DataFrame has columns like ['metric', 'value'] or similar,
    # try to detect metric/value columns
    candidates = [c.lower() for c in df.columns]
    if "metric" in candidates and ("value" in candidates or "val" in candidates):
        metric_col = df.columns[candidates.index("metric")]
        if "value" in candidates:
            value_col = df.columns[candidates.index("value")]
        else:
            value_col = df.columns[candidates.index("val")]
        return {str(r[metric_col]): r[value_col] for _, r in df.iterrows()}
    # fallback: if first column is metric and second is value
    if df.shape[1] >= 2:
        return {str(r[df.columns[0]]): r[df.columns[1]] for _, r in df.iterrows()}
    # otherwise nothing useful
    return {}

def gather_metrics_from_folder(folder_path: str, metrics_config: Dict[str, List[str]]) -> Dict[str, Any]:
    """
    Searches a given folder for the metrics files specificied in metrics_config, then extracts the relevant metrics.
    Returns a dictionary of extracted metrics flattened into dictionaries with single layer depth.

    """
    metric_dict: Dict[str, Any] = {}
    for metric_file, metric_info in metrics_config.items():
        #First read the file.
        table = pd.read_csv(os.path.join(folder_path,metric_file))

        for metric_name, extraction_info in metric_info.items():
            if extraction_info == None:
                #In this case, there should only be a single row. Lets verify that, then extract the relevant metric.
                if table.shape[0] != 1:
                    raise ValueError(f"Expected a single row in {metric_file} for metric {metric_name}, but found {table.shape[0]} rows. \n"
                                     "Please amend the metrics config to reflect the correct extraction strategy.")
                metric_dict[metric_name] = table.at[0, metric_name]
            else:
                #In this case, we have some extraction information to use, e.g. specific rows we need to extract.
                if 'iters' in extraction_info:
                    #We need to extract specific rows based on the iteration information provided. We will assume that 
                    #the first column contains the iteration information, but won't actually use this. Instead we will use
                    #string matching. 

                    for search_str in extraction_info['iters']:
                        mask = table.apply(lambda row: row.astype(str).str.contains(search_str, na=False)).any(axis=1)
                        indices = table.index[mask].tolist()
                        if len(indices) != 1:
                            raise ValueError(f"{len({indices})} rows found matching '{search_str}' in {metric_file} for metric {metric_name}. \n")
                        if indices[0] == 0:
                            metric_dict[f'{metric_name} Init.'] = table.at[indices[0], metric_name]
                        else:
                            metric_dict[f'{metric_name} Interactive Edit Iter {indices[0]}'] = table.at[indices[0], metric_name]
                else:
                    raise NotImplementedError("Unsupported extraction information provided in metrics config. Only iters is a supported \n" \
                    "configuration for multi-row extraction at the moment.")
            
        # metric_dict[metric_file] = metric_dict
    

    #Now we will reformat the dict so that the strings can be easily mapped to a table downstream.
    # reformatted_out = {v:out[k] for k,v in MAPPING_METRIC_CATEGORIES.items() if k in out}
    return metric_dict

def build_table(
    folders: dict[str],
    task_name: str,
    metrics_config: Dict[str, List[str]]) -> pd.DataFrame:
    """
    Create a DataFrame where rows are algorithms (by folder name) and columns are
    metric_type:metric_name (or just metric_name), in config order.
    
    inputs: 
    folders: dict mapping algorithm name to folder path
    metrics_config: dict mapping metric file names to configuration information for extraction.
    """

    metrics_storage = dict() 
    for folder, folder_path in folders.items():
        metrics_storage[folder] = gather_metrics_from_folder(folder_path, metrics_config)
        
    #Now using the extracted metrics to build a table with all of the relevant information placed together.


    #We will partition the table based on the metric type:
    #Dice, NSD, NoI. This is a fixed property for now. We will search through the dict of extracted metrics by creating a bank
    #of substrings to search for among each type.


    first_alg = list(metrics_storage.keys())[0]
    counts = dict()
    extracted_submetrics = dict()
    for k,v in RESULT_SEARCHSTRING.items():
        #We now count the number of submetrics for each metric type.
        #We do this by counting the number of items which have keys containing a substring matching that.

        #I.e., Dice_auc_median, or Dice_median. In particular, it should only flag for metrics which may have multiple sub
        #metrics that aren't distinct metric types. E.g., median dice at iter 0 and iter 100. 
        if isinstance(v, list):
            submetrics = [key for key in metrics_storage[first_alg] if any([substring in key for substring in v])]
        else:
            submetrics = [key for key in metrics_storage[first_alg] if v in key]

        counts[k] = len(submetrics)
        extracted_submetrics[k] = submetrics

    columns = [[k] * v for k,v in counts.items()]
    columns = [
        i
        for sublist in columns
        for i in sublist
    ]
    ordered_dict = {k:sort_by_substring_order(v, COLUMN_ORDERING) for k,v in extracted_submetrics.items()}
    subcolumns = [
        i
        for sublist in ordered_dict.values()
        for i in sublist
    ]
    #Now we create a multiindex for the columns.
    arrays = [
        columns,
        subcolumns
    ]
    tuples = list(zip(*arrays))
    flat_cols = [('Task', ''), ('Algorithm', '')]
    full_cols = pd.MultiIndex.from_tuples(flat_cols + tuples, names=["Metric Type", "Submetric"]) 
    rows = [] 
    for alg_name, metrics in metrics_storage.items():
        row = [task_name, alg_name]
        row += [metrics[col[-1]] if len(col) > 1 else metrics[col] for col in tuples]

        rows.append(row)
    df = pd.DataFrame(rows, columns=full_cols)
    
    # Round numeric columns to 3 decimal places
    if 'Failure_Cases_Fraction' in df.columns.get_level_values(1):
        #We need to convert this to a fraction.
        df[('NoI', 'Failure_Cases_Fraction')] = df[('NoI', 'Failure_Cases_Fraction')].to_numpy() / 100

    numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns
    df[numeric_cols] = df[numeric_cols].round(3)

    return df

def write_excel(df: pd.DataFrame, folder_root: str) -> None:
    """
    Write DataFrame to Excel. Overwrites if exists.
    """
    # Ensure parent dir exists
    os.makedirs(folder_root, exist_ok=True)
    df.to_excel(os.path.join(folder_root, 'summary.xlsx'), sheet_name="summary")

def write_csv(df: pd.DataFrame, folder_root: str) -> None:
    """
    Write DataFrame to CSV. Overwrites if exists.
    """
    os.makedirs(folder_root, exist_ok=True)
    df.to_csv(os.path.join(folder_root, 'summary.csv'), index=False)

    # Open the workbook and select the sheet
    # wb = load_workbook(os.path.join(folder_root, 'summary.xlsx'))
    # ws = wb.active

    # # Auto-adjust column width and enable wrap text
    # for col in ws.columns:
    #     max_length = 0
    #     col_letter = get_column_letter(col[0].column)
    #     for cell in col:
    #         try:
    #             if cell.value:
    #                 max_length = max(max_length, len(str(cell.value)))
    #         except:
    #             pass
    #     ws.column_dimensions[col_letter].width = min(max_length + 2, 50)  # limit max width
    #     for cell in col:
    #         cell.alignment = cell.alignment.copy(wrap_text=True)

    # wb.save(os.path.join(folder_root, 'summary.xlsx'))

def generate_latex(df: pd.DataFrame, caption: str = "", label: str = "", output_path: str = "") -> str:
    """
    Generate a LaTeX table string via pandas' to_latex.
    The returned string can be pasted into a LaTeX document.
    """

    # Generate LaTeX with MultiIndex columns
    df['Task'] = df['Task'].map(DATASET_MAPPING).fillna(df['Task'])
    df['Algorithm'] = df['Algorithm'].map(ALGORITHM_MAPPING).fillna(df['Algorithm'])
    latex_str = df.to_latex(index=False, multirow=True, multicolumn=True, longtable=False, caption=caption, label=label, na_rep="", float_format=lambda x: f'{x:g}')
    latex_str = merge_task_column_latex(latex_str, task_col=0)

    # Post-process LaTeX to merge 'Algorithm' header and remove subcolumn split for it
    lines = latex_str.splitlines()
    # Find the header line with column names (usually 4th line)
    for i, line in enumerate(lines):
        if '\\toprule' in line:
            header_start = i + 1
            lines[i] = '\hline' #We want a hline, not a toprule.
            break


    # The next two lines are the multiindex headers
    col_header = lines[header_start]
    subcol_header = lines[header_start + 1]
    
    #Lets tidy up this table header and the initial table object.
    subcol_number = df.columns.get_level_values(0).value_counts().to_dict() #This gives us a dict describing the number of subcolumns
    #per metric type. This is required to construct the table in the first place.
    tabular_shape = ''.join(['|' + 'c' * subcol_number[col_name] for col_name in df.columns.get_level_values(0).unique()]) + '|'
    table_env = [f'\\begin{{tabular}}{{{tabular_shape}}}', '\\hline']
    #First the task column.
    # col_header = ['\\multirow{2}{*}{Task}', '\\multirow{2}{*}{Algorithm}', '\\']
    col_header = [f'\\multirow{{2}}{{*}}{{{name}}}' if len(df[name].shape) == 1 else f'\\multicolumn{{{df[name].shape[1]}}}{{c|}}{{{name}}}' for name in df.columns.get_level_values(0).unique()]
    #Phew. This is a very involved line. Essentially we check if the column has subcolumns. If it does then we use multi-column structure. Otherwise
    #we use multirow structure. 
    
    #Now we need to construct the subcolumn header. Need to extract the indices of the final submetrics for each submetric
    #as this is required for the multicolumn structure.
    final_submetrics = {df[metric].columns[-1]:df.columns.get_level_values(1).get_loc(df[metric].columns[-1]) for metric in df.columns.get_level_values(0).unique() if len(df[metric].shape) > 1}
    subcol_header = ['' if submetric == '' else (f'\\multicolumn{{1}}{{c|}}{{{filter_string_table(submetric)}}}' if submetric not in final_submetrics else filter_string_table(submetric)) for submetric in df.columns.get_level_values(1)]

    # del lines[header_start - 2: header_start + 3] 
    lines[header_start - 2] = table_env[0]
    lines[header_start - 1] = table_env[1]
    lines[header_start] = ' & '.join(col_header) + ' \\\\ ' + f'\\cline{{3-{sum(subcol_number.values())}}}'
    lines[header_start + 1] = ' & '.join(subcol_header) + ' \\\\ \\hline'

    del lines[header_start + 2]
    #Now we filter the results rows. 

    for i, line in enumerate(lines):
        if '\\bottomrule' in line:
            result_end= i
            del lines[i] # lines[i] = '\hline' #We want a hline, not a toprule.
            break
    
    num_algo = len(df['Algorithm'].unique())
    #now filtering.
    for row_index, i in enumerate(range(header_start + 2, result_end)):
        split_row = lines[i].split('&')
        filtered_row = [substring if idx in final_submetrics.values()  or idx in [0,1] else f'\\multicolumn{{1}}{{c|}}{{{substring}}}' for idx, substring in enumerate(split_row)]
        if (row_index + 1) % num_algo:
            filtered_row.append('\\cline{{2-{}}}'.format(sum(subcol_number.values())))
        else:
            filtered_row.append('\\hline')
        lines[i] = ' & '.join(filtered_row[:-1]) + filtered_row[-1]
    
    # num_task = len(df['Task'].unique())
    
    # #Creating a duplicate list:
    # final_lines = lines[:header_start + 2]
    # #Now we will merge rows into single strings separated by &
    # for i in range(num_task):
    #     start_idx = header_start + 2 + i * num_algo
    #     end_idx = start_idx + num_algo
    #     merged_row = ''.join([lines[j] for j in range(start_idx, end_idx)])
    #     final_lines.append(merged_row)
    # final_lines.extend(lines[result_end:])

    # Add \usepackage{multirow} comment
    # final_lines.insert(0, '% Requires \\usepackage{multirow}')
    joined_lines = '\n'.join(lines)
    #We also need to add space for each & used.
    joined_lines = joined_lines.replace('&', '\n &')
    with open(output_path, 'w') as f:
        f.write(
            joined_lines
        )
    return

def merge_task_column_latex(latex_str, task_col=0):
    lines = latex_str.splitlines()
    start = next(i for i, l in enumerate(lines) if '\\midrule' in l) + 1
    end = next(i for i, l in enumerate(lines) if '\\bottomrule' in l)
    task_vals = [lines[i].split('&')[task_col].strip() for i in range(start, end)]
    i = 0
    while i < len(task_vals):
        val = task_vals[i]
        run = 1
        while i + run < len(task_vals) and task_vals[i + run] == val:
            run += 1
        if run > 1:
            lines[start + i] = lines[start + i].replace(val, f'\\multirow{{{run}}}{{*}}{{{val}}}', 1)
            for j in range(1, run):
                parts = lines[start + i + j].split('&')
                parts[task_col] = ' '
                lines[start + i + j] = '&'.join(parts)
        i += run
    lines.insert(0, '% Requires \\usepackage{multirow}')
    return '\n'.join(lines)

def parse_args():
    parser = argparse.ArgumentParser(description="Generate summary table (Excel + LaTeX) from algorithm result folders.")
    parser.add_argument("--metrics_root", type=str, required=True, help="Root path to the folder which contains the metrics across all the algorithms.")
    parser.add_argument("--algorithm_names", nargs="+", required=True, help="Names of algorithms to summarise.")
    #We explicitly pass algorithm names as we will use this to print the table!
    parser.add_argument("--experiment_subpath", nargs="+", required=True, help="Subpath under each algorithm folder where metrics are stored.")
    #NOTE: The subpath is assumed to be the same for all algorithms! 
    parser.add_argument("--metrics_config", required=True, help="JSON file mapping metrics to information required for pulling.")
    parser.add_argument("--output_root", required=True, help="Output root path for Excel (.xlsx) and LaTeX (.tex) files.")
    parser.add_argument("--caption", default="", help="Caption to include in LaTeX table.")
    parser.add_argument("--label", default="", help="Label to include in LaTeX table.")
    return parser.parse_args()

def main():
    args = parse_args()

    cfg = load_metrics_config(args.metrics_config)
    df_cumulative = pd.DataFrame()
    for experiment_subpath in args.experiment_subpath:
        folders = {alg_name: os.path.join(args.metrics_root, alg_name, experiment_subpath) for alg_name in args.algorithm_names}
        if os.name == 'posix':
            df_cumulative = pd.concat([df_cumulative, build_table(folders, experiment_subpath.split('/')[0], cfg)])
        else:
            raise NotImplementedError("Windows OS is not currently supported for table generation.")

    write_csv(df_cumulative, args.output_root)
    generate_latex(df_cumulative, caption=args.caption, label=args.label, output_path=args.output_root + '/summary_tex.txt')

if __name__ == "__main__":
    main()